import  openpyxl


wb = openpyxl.load_workbook("inventory.xlsx")
ws = wb["Sheet1"]

total_supplier_count = {}

for row in range(2, ws.max_row+1):
    supplier = ws.cell(row, 4).value

    total_supplier_count[supplier] = total_supplier_count.get(supplier) + 1

print(total_supplier_count)